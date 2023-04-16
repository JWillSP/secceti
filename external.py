from openpyxl import load_workbook
import io
from turmasdetalhes import turmastd, prodt


def delete_rows(ws, row):
  for i in range(row, ws.max_row + 1):
    ws.delete_rows(row)


def do_file(lista=[], turma='', file_name=''):

  wb = load_workbook(filename='./statics/' + file_name)
  ws = wb.active
  ws['F6'] = turma
  start_row = 11
  num_rows = start_row + len(lista)
  for i, item in enumerate(lista):
    row = start_row + i
    ws.cell(row=row, column=1).value = str(item)

  delete_rows(ws=ws, row=num_rows + 1)
  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data


def do_file2(lista=[], turma='', file_name='', prof='', comp='comp.: '):

  wb = load_workbook(filename='./statics/' + file_name)
  ws = wb.active
  turma_detail = turmastd.get(turma)
  turno = turma_detail.get('TURNO')
  if prof:
    nome_rm = prodt[1]
    matrícula = nome_rm.get(prof)
    pro_nome = prodt[0].get(matrícula)
  else:
    pro_nome = 'prof.: '
  ws['A3'] = turma + ' - ' + turno
  ws['B4'] = pro_nome
  ws['N4'] = comp
  start_row = 6
  num_rows = start_row + len(lista)
  for i, item in enumerate(lista):
    row = start_row + i
    ws.cell(row=row, column=1).value = str(item)

  delete_rows(ws=ws, row=num_rows + 1)
  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data
