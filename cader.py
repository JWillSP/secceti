from datetime import datetime
from openpyxl import load_workbook
import io
from turmasdetalhes import turmastd

#   if lfreq.shape[0] > 46:
#     source_filename = r'./static/Merged_fileVersionXL.xlsx'
# else:
#     source_filename = r'./static/Merged_file.xlsx'

source_filename = r'./statics/Merged_file.xlsx'
source_filename2 = r'./statics/cader_from_calldariaapp.xlsx'
def cf1(value):
    data = str(value).replace(',', '.')
    try:
        number = float(data)
    except ValueError:
        return 'FV'
    if 0 <= number <= 10:
        if number == 0: return '0'
        if number <= 3.5: return '3.5'
        if number <= 4.9: return '4.5'
        if number <= 10: return '7.5'
    else:
        return 'FV'

def normalizador(value):
    return {
        'FV': 'FV',
        '0': 'SC',
        '3.5': 'AC',
        '4.5': 'EC',
        '7.5': 'C'
    }.get(cf1(value), 'FV')

def do_cader(lfreq, turma, componente, professor, unidade):
  naulas_dadas = ''
  naulas = ''
  t_dict = turmastd.get(turma, {})
  turno = t_dict.get('TURNO')
  modalidade = t_dict.get('MODALIDADE')
  oferta_do_ensino = t_dict.get('OFERTA')
  serie = t_dict.get('ANO/SERIE/OUTROS')
  wb = load_workbook(filename=source_filename)
  ws1 = wb['diario']
  ws2 = wb['conteudo']

  ws1['A3'] = f'Ano/Série/Outros: {serie}'
  ws1['J2'] = f'Modalidade: {modalidade}'
  ws1['J3'] = f'Turma: {turma}'
  ws1['AH2'] = f'Oferta do Ensino: {oferta_do_ensino}'
  ws1['AH3'] = f'Turno: {turno}'
  ws1['A4'] = f'Componente: {componente}'
  ws1['A5'] = f'Unidade: {unidade}'
  ws1['J4'] = f'Professor: {professor}'
  ws1['J5'] = f'Aulas Previstas: {naulas}'
  ws1['AH5'] = f'Aulas Dadas: {naulas_dadas}'

  for row, nome in enumerate(lfreq):
    ws1[F'B{row + 9}'] = nome

  ws2['A3'] = f'Ano/Série/Outros: {serie}'
  ws2['J2'] = f'Modalidade: {modalidade}'
  ws2['J3'] = f'Turma: {turma}'
  ws2['AH2'] = f'Oferta do Ensino: {oferta_do_ensino}'
  ws2['AH3'] = f'Turno: {turno}'
  ws2['A4'] = f'Componente: {componente}'
  ws2['A5'] = f'Unidade: {unidade}'
  ws2['J4'] = f'Professor: {professor}'
  ws2['J5'] = f'Aulas Previstas: {naulas}'
  ws2['AH5'] = f'Aulas Dadas: {naulas_dadas}'

  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data


def do_cader_x(
    lmed, turma,
    componente,
    professor,
    unidade,
    aula_id={},
    aula_data={},
    aula_content_dict={},
    pivot=None,
    pivot2=None
):
  naulas_dadas = ''
  naulas = ''
  t_dict = turmastd.get(turma, {})
  turno = t_dict.get('TURNO')
  modalidade = t_dict.get('MODALIDADE')
  oferta_do_ensino = t_dict.get('OFERTA')
  serie = t_dict.get('ANO/SERIE/OUTROS')
  is_eja = 'TF' in turma or 'TJ' in turma
  wb = load_workbook(filename=source_filename2)
  ws1 = wb['diario']
  ws2 = wb['conteudo']

  timestamps = list(aula_data.items())
  sorted_timestamps = sorted(timestamps, key=lambda x: datetime.timestamp(x[1]))
  for index, timestamp in enumerate(sorted_timestamps):
    ws1.cell(row=6, column=(3 + index), value=timestamp[0])
  ws1['A3'] = f'Ano/Série/Outros: {serie}'
  ws1['J2'] = f'Modalidade: {modalidade}'
  ws1['J3'] = f'Turma: {turma}'
  ws1['AH2'] = f'Oferta do Ensino: {oferta_do_ensino}'
  ws1['AH3'] = f'Turno: {turno}'
  ws1['A4'] = f'Componente: {componente}'
  ws1['A5'] = f'Unidade: {unidade}'
  ws1['J4'] = f'Professor: {professor}'
  ws1['J5'] = f'Aulas Previstas: {naulas}'
  ws1['AH5'] = f'Aulas Dadas: {naulas_dadas}'

  for row, std_obj in enumerate(lmed):
    rm = std_obj.get('matrícula')
    ws1[F'B{row + 9}'] = std_obj.get('name')
    if is_eja:
      ws1[F'BA{row + 9}'] = normalizador(std_obj.get('med'))
    else:
      ws1[F'BA{row + 9}'] = std_obj.get('med')
    if not pivot2.empty:
      ws1[F'BB{row + 9}'] = pivot.loc[:, 'TOTAL_DE_FALTAS'].to_dict().get(rm, '0')
    for index, timestamp in enumerate(sorted_timestamps):
      if timestamp[0] in pivot.columns:
        jus = pivot[timestamp[0]].to_dict().get(rm, 0) - pivot2[timestamp[0]].to_dict().get(rm, 0)
        if jus:
          ws1.cell(row=(row + 9), column=(3 + index), value=f'{jus}J')
        else:
          faltas_na_data = pivot[timestamp[0]].to_dict().get(rm, '0')
          to_fill = str(faltas_na_data).replace('0', '∙').replace('1', '1F').replace('2', '2F')
          ws1.cell(row=(row + 9), column=(3 + index), value=to_fill)
      else:
        ws1.cell(row=(row + 9), column=(3 + index), value='∙')
  for index, timestamp in enumerate(sorted_timestamps):
    ws2[f'B{index + 9}'] = timestamp[0]
    ws2[f'C{index + 9}'] = aula_content_dict.get(timestamp[0], '')
    ws2[f'AH{index + 9}'] = aula_id.get(timestamp[0], '')

  ws2['A3'] = f'Ano/Série/Outros: {serie}'
  ws2['J2'] = f'Modalidade: {modalidade}'
  ws2['J3'] = f'Turma: {turma}'
  ws2['AH2'] = f'Oferta do Ensino: {oferta_do_ensino}'
  ws2['AH3'] = f'Turno: {turno}'
  ws2['A4'] = f'Componente: {componente}'
  ws2['A5'] = f'Unidade: {unidade}'
  ws2['J4'] = f'Professor: {professor}'
  ws2['J5'] = f'Aulas Previstas: {naulas}'
  ws2['AH5'] = f'Aulas Dadas: {naulas_dadas}'

  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data